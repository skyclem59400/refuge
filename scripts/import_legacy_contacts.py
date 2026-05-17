#!/usr/bin/env python3
"""
Importe les contacts d'un export Hunimalis XLSX dans la table legacy_contacts.

Usage:
    python3 scripts/import_legacy_contacts.py <chemin_xlsx> [--establishment-id <uuid>] [--source hunimalis_2026]

Le script :
  1. Lit le fichier XLSX (5 colonnes : Nom, Adresse, Code postal, Ville, Tél)
  2. Normalise nom (Title Case), nom_normalized (lowercase sans accents),
     téléphone (E.164 si reconnaissable comme FR)
  3. Déduplique intra-fichier sur (full_name_normalized, phone_normalized, postal_code)
  4. Insère par batches de 500 via l'API REST Supabase
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl manquant : pip3 install --user openpyxl")

try:
    import requests
except ImportError:
    sys.exit("requests manquant : pip3 install --user requests")


SDA_ESTAB_ID = "f0a9d4a8-143d-431e-a875-0b2dc1f505ba"
BATCH_SIZE = 500


def load_env(env_path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not env_path.exists():
        return env
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def title_case_name(raw: str) -> str:
    """Normalise un nom : ABRAHAM JEROME → Abraham Jerome, abdelmajid ali → Abdelmajid Ali."""
    parts = re.split(r"(\s+|-|'|/)", raw.strip())
    return "".join(p.capitalize() if p.strip() and p not in {"-", "'", "/"} else p for p in parts)


def normalize_name_for_search(name: str) -> str:
    """Lowercase + strip accents + collapse spaces — pour recherche trigram."""
    s = strip_accents(name).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_phone(raw: Any) -> str | None:
    """
    Normalise un téléphone FR :
      - +33781361943 → +33781361943
      - 766789973    → +33766789973 (9 chiffres = mobile sans le 0)
      - 0601775072   → +33601775072
      - sinon retourne le brut strippé ou None
    """
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    s = re.sub(r"[^\d+]", "", s)
    if not s:
        return None
    if s.startswith("+33") and len(s) == 12:
        return s
    if s.startswith("33") and len(s) == 11:
        return "+" + s
    if s.startswith("0") and len(s) == 10:
        return "+33" + s[1:]
    if len(s) == 9 and s[0] in "67":
        return "+33" + s
    return s if s else None


def is_placeholder_name(name: str) -> bool:
    stripped = name.strip()
    if not stripped:
        return True
    if stripped in {"...", "... ...", ".... ....", ".........", "?", "??", "???"}:
        return True
    # noms commençant par "??" suivis de quelque chose : on garde (ex: "?? Corinne")
    # noms uniquement constitués de ponctuation
    if not re.search(r"[a-zA-ZÀ-ÿ]", stripped):
        return True
    return False


def clean_text(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def clean_postal(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    # Postal FR = 5 chiffres. Le XLSX peut avoir des int (62117) ou str.
    digits = re.sub(r"\D", "", s)
    if 4 <= len(digits) <= 5:
        return digits.zfill(5)
    return s if s else None


def parse_xlsx(path: Path) -> list[dict[str, Any]]:
    print(f"Lecture {path.name}...", flush=True)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    seen_keys: set[tuple] = set()
    contacts: list[dict[str, Any]] = []
    dropped_placeholders = 0
    dropped_dup_in_file = 0

    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not any(r):
            continue
        nom_raw = clean_text(r[0] if len(r) > 0 else None)
        if not nom_raw or is_placeholder_name(nom_raw):
            dropped_placeholders += 1
            continue

        full_name = title_case_name(nom_raw)
        full_name_normalized = normalize_name_for_search(full_name)
        address = clean_text(r[1] if len(r) > 1 else None)
        postal_code = clean_postal(r[2] if len(r) > 2 else None)
        city = clean_text(r[3] if len(r) > 3 else None)
        if city:
            city = title_case_name(city)
        phone_raw = clean_text(r[4] if len(r) > 4 else None)
        phone_normalized = normalize_phone(phone_raw)

        dedup_key = (full_name_normalized, phone_normalized or "", postal_code or "")
        if dedup_key in seen_keys:
            dropped_dup_in_file += 1
            continue
        seen_keys.add(dedup_key)

        contacts.append({
            "establishment_id": SDA_ESTAB_ID,
            "source": "hunimalis_2026",
            "full_name": full_name,
            "full_name_normalized": full_name_normalized,
            "address": address,
            "postal_code": postal_code,
            "city": city,
            "phone": phone_raw,
            "phone_normalized": phone_normalized,
        })

    print(f"  → {len(contacts):,} contacts retenus")
    print(f"  → {dropped_placeholders:,} lignes placeholder/vides ignorées")
    print(f"  → {dropped_dup_in_file:,} doublons intra-fichier dédupliqués")
    return contacts


def insert_batches(contacts: list[dict[str, Any]], supabase_url: str, service_key: str) -> None:
    endpoint = f"{supabase_url}/rest/v1/legacy_contacts"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    total = len(contacts)
    inserted = 0
    failures = 0
    start = time.time()

    for i in range(0, total, BATCH_SIZE):
        batch = contacts[i:i + BATCH_SIZE]
        try:
            resp = requests.post(endpoint, json=batch, headers=headers, timeout=60)
            if resp.status_code >= 300:
                failures += len(batch)
                print(f"  ✗ batch {i // BATCH_SIZE + 1} échoué : HTTP {resp.status_code} — {resp.text[:200]}")
            else:
                inserted += len(batch)
        except requests.RequestException as e:
            failures += len(batch)
            print(f"  ✗ batch {i // BATCH_SIZE + 1} exception : {e}")

        if (i // BATCH_SIZE + 1) % 10 == 0 or (i + BATCH_SIZE) >= total:
            elapsed = time.time() - start
            print(f"  {inserted:,}/{total:,} insérés ({elapsed:.1f}s)", flush=True)

    print(f"\n✓ Import terminé : {inserted:,} insérés, {failures} échecs en {time.time() - start:.1f}s")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path, help="Chemin du fichier XLSX Hunimalis")
    parser.add_argument("--establishment-id", default=SDA_ESTAB_ID, help="UUID de l'établissement (défaut SDA)")
    parser.add_argument("--dry-run", action="store_true", help="Parse sans insérer")
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        sys.exit(f"Fichier introuvable : {args.xlsx_path}")

    env = load_env(Path(__file__).parent.parent / ".env.local")
    supabase_url = env.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        sys.exit("NEXT_PUBLIC_SUPABASE_URL et SUPABASE_SERVICE_ROLE_KEY requis (dans .env.local ou env)")

    contacts = parse_xlsx(args.xlsx_path)
    if not contacts:
        sys.exit("Aucun contact à importer.")

    if args.dry_run:
        print("DRY RUN — pas d'insertion. Aperçu des 3 premiers :")
        for c in contacts[:3]:
            print(f"  {c}")
        return 0

    # Override establishment_id si spécifié
    if args.establishment_id != SDA_ESTAB_ID:
        for c in contacts:
            c["establishment_id"] = args.establishment_id

    print(f"\nInsertion par batches de {BATCH_SIZE} dans Supabase...")
    insert_batches(contacts, supabase_url, service_key)
    return 0


if __name__ == "__main__":
    sys.exit(main())
